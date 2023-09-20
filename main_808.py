import sys
import os
import time
import serial
import datetime
import openpyxl
import webbrowser
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from ui_808 import Ui_Form
from ui_front import ui_front_v1
from fun_lib import creat_excel
from fun_lib import wgs84_to_gcj02
from BDS_plot_function import BDS_plot
from BDS_distance_function import BDS_distance
from deleter_fun import del_xlsx
# from LL2city_720_ok import LL2city_name


# 613 1.速度输出修改至线程一
# 618 1.使用ui版本618为操作界面 2.增加大豆株距传入 3.修改株据传入为玉米株距传入 4.串口模式增加usb模式
# 705 1.给株据交互模块加入串口清空 2.加入株据信息交互暂停时，GPS信息交互同步暂停的功能
# 709 1.株据交互加入判断符
# 715 1.加入欢迎界面
# 719 1.通讯加入报错继续程序：a.GNSS数学错误 b.通讯模块串口异常
# 720 1.加入地理信息逆编码 2.加入线程3负责上位机与种箱监测模块通信
# 808 1.加入测速雷达测速模块 2.GNSS速度信息与测速雷达速度信息融合 3.使用ui_808

class Worker_1(QThread):
    '''
    线程一负责上位机与GPS模块通讯
    '''
    finished_signal = pyqtSignal()
    tran_speed = pyqtSignal(float)
    # city_name_transmit = pyqtSignal()
    mistake_message_transmit = pyqtSignal(str)

    def __init__(self, port):
        super().__init__()
        self.now = None
        self.name = None
        self.should_run = True
        self.port = port
        self.speed = 0
        self.gnss_message = 'work'

    def set_gnss_pause_message(self, message):
        self.gnss_message = message
        pass

    def run(self):
        with serial.Serial(self.port, 9600, timeout=1) as ser:
            i = 1
            rip = 0
            # 获取当前时间
            self.now = datetime.datetime.now()
            self.name = self.now.strftime("%Y%m%d%H%M")
            creat_excel(self.name)
            while self.should_run:
                data = ser.readline().decode().strip()
                try:
                    if data == "INVALID":
                        mistake = '无信号，请重置GNSS模块'
                        self.mistake_message_transmit.emit(str(mistake))
                        print(mistake)
                        continue
                        pass
                    elif data == '':
                        print('信号接收丢失，请重置GNSS模块')
                        continue
                        pass
                    else:
                        nums = [float(num_str) for num_str in data.split(",")]
                        x = nums[0]
                        y = nums[1]
                        h = nums[2]
                        v = nums[3]
                        x, y = wgs84_to_gcj02(x, y)
                        # 数据合理性判断
                        if 73 < x < 135 and 18 < y < 55 and h >= 0 and v >= 0 and self.gnss_message == 'work':
                            global V
                            V = v
                            name = self.name + '.xlsx'
                            workbook = openpyxl.load_workbook(name)
                            sheet = workbook.active
                            j = 1
                            sheet.cell(row=i, column=j, value=x)
                            sheet.cell(row=i, column=j + 1, value=y)
                            sheet.cell(row=i, column=j + 2, value=h)
                            sheet.cell(row=i, column=j + 3, value=v)
                            i = i + 1
                            workbook.save(name)
                            self.tran_speed.emit(V)
                            # if rip == 0:
                            #     city_name = LL2city_name(y, x)
                            #     self.city_name_transmit.emit(city_name)
                            #     rip = 1
                            #     pass
                            print(x, y, h, v)
                            pass
                        elif 3 < x < 135 and 18 < y < 55 and h >= 0 and v >= 0 and self.gnss_message == 'pause':
                            mistake = '已暂停记录'
                            self.mistake_message_transmit.emit(str(mistake))
                            print(mistake)
                        else:
                            print('信号数据有误，请等待')
                            continue
                            pass
                        pass
                    pass
                except ValueError:
                    continue
                    pass
                except Exception as err_1:
                    self.mistake_message_transmit.emit(str(err_1))
                pass
            pass
        self.finished_signal.emit()
        pass


class Worker_2(QThread):
    '''
    线程二负责上位机与播种单元的控制
    '''
    finished_signal = pyqtSignal()

    def __init__(self, port, spacing_YM, spacing_DD):
        super().__init__()
        self.should_run = True
        self.port = port
        self.Spacing_YM = spacing_YM
        self.Spacing_DD = spacing_DD
        self._message = 'OK'
        self.signal = 1
        self.ser = serial.Serial(self.port, 115200, timeout=1)
        pass

    def set_message(self, message):
        self._message = message
        pass

    def run(self):
        while self.should_run:
            if self._message == "pause":
                self.signal = 1
                pass
            elif self._message == 'OK':
                self.signal = 0
                pass
            A = str(self.signal) + ',' + str(self.Spacing_YM) + ',' + str(self.Spacing_DD) + ',' + str(V)
            print(A)
            self.ser.write(str(A).encode() + b'\n')
            self.ser.flushInput()
            self.ser.flushOutput()
            time.sleep(0.5)
        self.finished_signal.emit()
        pass

    pass


class Worker_3(QThread):
    '''
    线程三负责种箱检测和预警
    '''
    finished_signal = pyqtSignal()
    send_ratio_one = pyqtSignal(int)
    send_ratio_two = pyqtSignal(int)

    def __init__(self, port, box_value):
        super(Worker_3, self).__init__()
        self.should_run = True
        self.port = port
        self._message = 'OK'
        self.ser = serial.Serial(self.port, 19200, timeout=1)
        self.box_value = box_value
        pass

    def set_message(self, message):
        self._message = message
        pass

    def run(self):
        self.ser.write(str(self.box_value).encode() + b'\n')
        self.box_value = self.box_value * 100
        while self.should_run:
            # 增加适当的延时，确保数据完全接收
            data = self.ser.readline().decode().strip()
            print(data)
            try:
                Identifier, Type, Data = data.split(',')
                Identifier = int(Identifier)
                Type = int(Type)
                Data = float(Data)
                if Identifier == 0:
                    if Type == 1 and self._message == 'OK':
                        ratio = int(((self.box_value - Data) / self.box_value) * 100)
                        self.send_ratio_one.emit(ratio)
                        pass
                    elif Type == 2 and self._message == 'OK':
                        ratio = int(((self.box_value - Data) / self.box_value) * 100)
                        self.send_ratio_two.emit(ratio)
                        pass
                    elif Type != 1 or Type != 2:
                        print('数据有误，无法区分种箱编码')
                        pass
                    else:
                        print('其他错误')
                        pass
                    pass
                else:
                    print('数据有误，无法工作')
                    continue
                    pass
                pass
            except ValueError:
                continue
                pass
            except Exception as err_1:
                print("新错误：", err_1)
                pass
            pass
        self.finished_signal.emit()
        pass

    pass


class Worker_4(QThread):
    '''
    线程4负责测速雷达测速
    '''
    send_speed_information = pyqtSignal(float)
    finished_signal = pyqtSignal()
    show_speed_for_ratio = pyqtSignal()

    def __init__(self, port):
        super(Worker_4, self).__init__()
        self.should_run = True
        self.port = port
        self.ser = serial.Serial(self.port, 38400, timeout=1)
        pass

    def run(self):
        A = 1
        while self.should_run:
            self.ser.write(str(A).encode() + b'\n')
            data = self.ser.readline().decode().strip()
            print(data)
            try:
                identifier, speed = data.split(',')
                identifier = int(identifier)
                speed = float(speed)
                if identifier == 1:
                    if speed == float('inf'):
                        speed = 0
                        pass
                    self.send_speed_information.emit(speed)
                    self.show_speed_for_ratio.emit()
                    pass
            except ValueError:
                continue
                pass
            except Exception as err_1:
                print("新错误：", err_1)
                pass
            pass
        self.finished_signal.emit()
        pass


class MyWindow(QWidget):
    def __init__(self, port_one, port_two, port_three, port_four):
        super(MyWindow, self).__init__()
        self.radar_speed = None
        self.gnss_speed = 0
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.worker_one = None
        self.worker_two = None
        self.worker_third = None
        self.worker_four = None
        self.port_one = port_one
        self.port_two = port_two
        self.port_three = port_three
        self.port_four = port_four
        self.ui_process()

    # UI模组
    def ui_process(self):
        # 读取ui控件
        # GPS模块
        self.accept = self.ui.pushButton6
        self.end = self.ui.pushButton5
        # 株距控制模块
        self.Spacing_begin = self.ui.pushButton7
        self.Spacing_over = self.ui.pushButton8
        self.Spacing_control_YM = self.ui.spinBox
        self.Spacing_control_YM.setValue(0)
        self.Spacing_control_DD = self.ui.spinBox_2
        self.Spacing_control_DD.setValue(0)
        self.pause = self.ui.pushButton9
        # 速度显示模块
        self.lcd = self.ui.lcdNumber
        # 后处理模块
        self.openfile = self.ui.pushButton1
        self.make_map = self.ui.pushButton2
        self.calculate_distance = self.ui.pushButton3
        self.deleter_file = self.ui.pushButton4
        self.textBrowser = self.ui.textBrowser
        self.width_control = self.ui.doubleSpinBox
        self.width_control.setValue(0.0)
        # 串口设置模块
        self.ui.comboBox.addItem('ACM0')
        self.ui.comboBox.addItem('ACM1')
        self.ui.comboBox.addItem('ACM2')
        self.ui.comboBox.addItem('USB0')
        self.ui.comboBox.addItem('USB1')
        self.ui.comboBox_2.addItem('ACM0')
        self.ui.comboBox_2.addItem('ACM1')
        self.ui.comboBox_2.addItem('ACM2')
        self.ui.comboBox_2.addItem('USB0')
        self.ui.comboBox_2.addItem('USB1')
        self.ui.comboBox_3.addItem('ACM0')
        self.ui.comboBox_3.addItem('ACM1')
        self.ui.comboBox_3.addItem('ACM2')
        self.ui.comboBox_3.addItem('USB0')
        self.ui.comboBox_3.addItem('USB1')
        self.ui.comboBox_4.addItem('ACM0')
        self.ui.comboBox_4.addItem('ACM1')
        self.ui.comboBox_4.addItem('ACM2')
        self.ui.comboBox_4.addItem('USB0')
        self.ui.comboBox_4.addItem('USB1')
        # 种箱监测模块
        self.box_monitor_begin = self.ui.pushButton7_2
        self.box_monitor_end = self.ui.pushButton8_2
        self.zhongxiang_1 = self.ui.progressBar
        self.zhongxiang_2 = self.ui.progressBar_2
        self.zhongxiang_1.setValue(100)
        self.zhongxiang_2.setValue(100)
        self.zhongxiangfuwei_1 = self.ui.pushButton
        self.zhongxiangfuwei_2 = self.ui.pushButton_2
        self.box_value = self.ui.doubleSpinBox_2
        self.box_value.setValue(0.3)
        # 绑定信号与函数
        # GPS模块
        self.accept.clicked.connect(self.information_present)
        self.end.clicked.connect(self.information_end)
        self.ui.comboBox.currentIndexChanged.connect(self.choose_port_1)
        # 后处理模块
        self.openfile.clicked.connect(self.open_file)
        self.make_map.clicked.connect(self.plot_map)
        self.make_map.clicked.connect(self.pop)
        self.deleter_file.clicked.connect(self.del_file)
        self.calculate_distance.clicked.connect(self.distance)
        # 株距控制模块
        self.Spacing_begin.clicked.connect(self.Spacing_present)
        self.Spacing_over.clicked.connect(self.Spacing_end)
        self.ui.comboBox_2.currentIndexChanged.connect(self.choose_port_2)
        self.pause.clicked.connect(self.pause_process)
        # 种箱监测模块
        self.box_monitor_begin.clicked.connect(self.seed_box_information_present)
        self.box_monitor_end.clicked.connect(self.seed_box_information_end)
        self.ui.comboBox_3.currentIndexChanged.connect(self.choose_port_3)
        self.zhongxiangfuwei_1.clicked.connect(self.zhongxiangfuwei_process_one)
        self.zhongxiangfuwei_2.clicked.connect(self.zhongxiangfuwei_process_two)
        # 测速雷达模块
        self.ui.comboBox_4.currentIndexChanged.connect(self.choose_port_4)
        pass

    # GPS模块串口选择
    def choose_port_1(self):
        if self.ui.comboBox.currentText() == 'ACM0':
            self.port_one = "/dev/ttyACM0"
            pass
        elif self.ui.comboBox.currentText() == 'ACM1':
            self.port_one = "/dev/ttyACM1"
            pass
        elif self.ui.comboBox.currentText() == 'USB0':
            self.port_one = "/dev/ttyUSB0"
            pass
        elif self.ui.comboBox.currentText() == 'USB1':
            self.port_one = "/dev/ttyUSB1"
            pass
        elif self.ui.comboBox.currentText() == 'ACM2':
            self.port_one = "/dev/ttyACM2"
            pass
        pass

    # 株距控制模块串口选择
    def choose_port_2(self):
        if self.ui.comboBox_2.currentText() == 'ACM0':
            self.port_two = "/dev/ttyACM0"
            pass
        elif self.ui.comboBox_2.currentText() == 'ACM1':
            self.port_two = "/dev/ttyACM1"
            pass
        elif self.ui.comboBox_2.currentText() == 'USB0':
            self.port_two = "/dev/ttyUSB0"
            pass
        elif self.ui.comboBox_2.currentText() == 'USB1':
            self.port_two = "/dev/ttyUSB1"
            pass
        elif self.ui.comboBox_2.currentText() == 'ACM2':
            self.port_two = "/dev/ttyACM2"
            pass
        pass

    # 种箱监测模块串口选择
    def choose_port_3(self):
        if self.ui.comboBox_3.currentText() == 'ACM0':
            self.port_three = "/dev/ttyACM0"
            pass
        elif self.ui.comboBox_3.currentText() == 'ACM1':
            self.port_three = "/dev/ttyACM1"
            pass
        elif self.ui.comboBox_3.currentText() == 'USB0':
            self.port_three = "/dev/ttyUSB0"
            pass
        elif self.ui.comboBox_3.currentText() == 'USB1':
            self.port_three = "/dev/ttyUSB1"
            pass
        elif self.ui.comboBox_3.currentText() == 'ACM2':
            self.port_three = "/dev/ttyACM2"
            pass
        pass

    # 测速雷达串口选择
    def choose_port_4(self):
        if self.ui.comboBox_4.currentText() == 'ACM0':
            self.port_four = "/dev/ttyACM0"
            pass
        elif self.ui.comboBox_4.currentText() == 'ACM1':
            self.port_four = "/dev/ttyACM1"
            pass
        elif self.ui.comboBox_4.currentText() == 'USB0':
            self.port_four = "/dev/ttyUSB0"
            pass
        elif self.ui.comboBox_4.currentText() == 'USB1':
            self.port_four = "/dev/ttyUSB1"
            pass
        elif self.ui.comboBox_4.currentText() == 'ACM2':
            self.port_four = "/dev/ttyACM2"
            pass
        pass

    # GPS模块通讯
    def information_present(self):
        try:
            self.worker_one = Worker_1(self.port_one)
            self.worker_one.finished_signal.connect(self.worker_finished)
            self.worker_one.tran_speed.connect(self.speed_gnss)
            # self.worker_one.city_name_transmit(self.show_city_name)
            self.worker_one.mistake_message_transmit.connect(self.mistake_message_show)
            self.worker_one.start()
            pass
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(error)
            pass
        pass

    def information_end(self):
        if self.worker_one is not None:
            self.worker_one.should_run = False
        pass

    def worker_finished(self):
        self.worker_one = None
        pass

    def show_city_name(self, city_name):
        self.textBrowser.setText("欢迎来自{}的朋友".format(city_name))
        pass

    # 种箱监测模块
    def seed_box_information_present(self):
        box_value = self.box_value.value()
        try:
            self.worker_third = Worker_3(self.port_three, box_value)
            self.worker_third.finished_signal.connect(self.seed_box_information_end)
            self.worker_third.send_ratio_one.connect(self.show_ratio_one)
            self.worker_third.send_ratio_two.connect(self.show_ratio_two)
            self.worker_third.start()
            pass
        except serial.serialutil.SerialException as error:
            print('窗口打开无效，请重新调整串口', error)
            self.mistake_message_show(error)
            pass
        pass

    def show_ratio_one(self, ratio):
        self.zhongxiang_1.setValue(ratio)
        pass

    def show_ratio_two(self, ratio):
        self.zhongxiang_2.setValue(ratio)
        pass

    def seed_box_information_end(self):
        if self.worker_third is not None:
            self.worker_third.should_run = False
            pass
        pass

    def worker_third_finished(self):
        self.worker_third = None
        pass

    def zhongxiangfuwei_process_one(self):
        self.worker_third.set_message('OK')
        self.zhongxiang_1.setValue(100)
        pass

    def zhongxiangfuwei_process_two(self):
        self.worker_third.set_message('OK')
        self.zhongxiang_2.setValue(100)
        pass

    def pop_two(self, which):
        if which == 1:
            QMessageBox.information(self, "确认", "种箱1种料不足")
            pass
        if which == 2:
            QMessageBox.information(self, "确认", "种箱2种料不足")
            pass
        pass

    # 后处理模块
    def open_file(self):
        self.filename = QFileDialog.getOpenFileNames(self, '选择文件', os.getcwd(), "All Files(*)")
        self.path = self.filename[0]
        pass

    def plot_map(self):
        BDS_plot(self.path[0])
        webbrowser.open_new_tab('loc_map.html')
        pass

    def distance(self):
        n = self.width_control.value()
        m = BDS_distance(self.path[0], n)
        self.textBrowser.setText("工作距离为{}米,工作面积为{}平方米".format(m[0], m[1]))
        self.textBrowser.repaint()
        pass

    def pop(self):
        QMessageBox.information(self, "确认", "地图生成")
        pass

    def del_file(self):
        self.directory = QFileDialog.getExistingDirectory(self, '选择文件夹', "./")
        del_xlsx(self.directory)
        pass

    # 株距控制模块
    '''
    株距控制模块也负责测速雷达的启动
    '''
    def Spacing_present(self):
        YM = self.Spacing_control_YM.value()
        DD = self.Spacing_control_DD.value()
        try:
            self.worker_two = Worker_2(self.port_two, YM, DD)
            self.worker_two.finished_signal.connect(self.worker_two_finished)
            self.worker_two.start()
            self.worker_four = Worker_4(self.port_four)
            self.worker_four.start()
            self.worker_four.finished_signal.connect(self.worker_four_finished)
            self.worker_four.send_speed_information.connect(self.speed_radar)
            self.worker_four.show_speed_for_ratio.connect(self.show_speed)
            pass
        except serial.serialutil.SerialException as err:
            self.mistake_message_show(err)
        pass

    def Spacing_end(self):
        if self.worker_two is not None:
            self.worker_two.should_run = False
            self.worker_four.should_run = False
        pass

    def worker_two_finished(self):
        self.worker_two = None
        pass

    def worker_four_finished(self):
        self.worker_four = None
        pass

    def pause_process(self):
        if self.worker_two._message == "OK":
            self.worker_two.set_message("pause")
            self.worker_one.set_gnss_pause_message('pause')
            pass
        else:
            self.worker_two.set_message("OK")
            self.worker_one.set_gnss_pause_message('work')
        pass

    # 速度显示模块
    def speed_gnss(self, speed):
        self.gnss_speed = speed
        pass

    def speed_radar(self, speed):
        self.radar_speed = speed
        pass

    def show_speed(self):
        if self.radar_speed == 0:
            speed = 0
            pass
        else:
            # 比例初步确定为3：7
            speed = self.gnss_speed * 0.3 + self.radar_speed * 0.7
            pass
        self.lcd.display(speed)
        self.lcd.show()
        pass

    # 错误信息显示模块
    def mistake_message_show(self, message):
        QMessageBox.information(self, "错误", "错误原因{}".format(message))


class ui_front(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = ui_front_v1()
        self.ui.setupUi(self)
        self.ui.pushButton.clicked.connect(self.switch_window)
        pass

    def switch_window(self):
        stacked_widget.setCurrentIndex(1)
        pass


if __name__ == '__main__':
    port_one = "/dev/ttyACM0"
    port_two = "/dev/ttyACM0"
    port_three = "/dev/ttyACM0"
    port_four = "/dev/ttyACM0"
    V = 5
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    app = QApplication(sys.argv)
    stacked_widget = QStackedWidget()
    window = MyWindow(port_one, port_two, port_three, port_four)
    ui_front_window = ui_front()
    stacked_widget.addWidget(ui_front_window)
    stacked_widget.addWidget(window)
    stacked_widget.setWindowTitle("My Application")
    stacked_widget.setCurrentIndex(0)
    design_widget = QWidget()
    design_window = Ui_Form()
    design_window.setupUi(design_widget)
    design_size = design_widget.size()
    stacked_widget.setGeometry(0, 0, design_size.width(), design_size.height())
    # 显示堆叠窗口
    stacked_widget.show()
    sys.exit(app.exec_())
    pass
